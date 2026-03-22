#!/usr/bin/env python3
"""
generate_sequences.py — S6 dv-sequences generator
Generates UVM sequence and test class files from a testplan + TB data JSON.

Usage:
  python3 generate_sequences.py --input /tmp/<proj>_sequences_input.json --output <dv_root>
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path
from datetime import date

# ── openpyxl (optional — needed only when reading xlsx directly) ──────────────
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Protocol signal databases ─────────────────────────────────────────────────
PROTOCOL_SIGNALS = {
    "APB": ["pclk", "presetn", "paddr", "psel", "penable", "pwrite", "pwdata",
            "prdata", "pready", "pslverr"],
    "AHB": ["hclk", "hresetn", "haddr", "htrans", "hwrite", "hsize", "hburst",
            "hprot", "hwdata", "hrdata", "hready", "hresp"],
    "AHB-Lite": ["hclk", "hresetn", "haddr", "htrans", "hwrite", "hsize",
                 "hwdata", "hrdata", "hreadyout", "hresp"],
    "AXI4": ["aclk", "aresetn", "awaddr", "awlen", "awsize", "awburst", "awvalid",
             "awready", "wdata", "wstrb", "wlast", "wvalid", "wready",
             "bresp", "bvalid", "bready", "araddr", "arlen", "arsize", "arburst",
             "arvalid", "arready", "rdata", "rresp", "rlast", "rvalid", "rready"],
    "AXI4-Stream": ["aclk", "aresetn", "tvalid", "tready", "tdata", "tstrb",
                    "tkeep", "tlast", "tid", "tdest", "tuser"],
    "SPI": ["sclk", "cs_n", "mosi", "miso"],
    "I2C": ["scl", "sda"],
    "UART": ["clk", "rst_n", "txd", "rxd", "rts_n", "cts_n"],
    "TileLink": ["clock", "reset", "a_opcode", "a_param", "a_size", "a_source",
                 "a_address", "a_mask", "a_data", "a_valid", "a_ready",
                 "d_opcode", "d_param", "d_size", "d_source", "d_sink",
                 "d_data", "d_error", "d_valid", "d_ready"],
}

# Operations per protocol → used to decide which seq files to generate
PROTOCOL_OPERATIONS = {
    "APB":        ["write", "read", "poll"],
    "AHB":        ["write", "read", "burst_write", "burst_read"],
    "AHB-Lite":   ["write", "read", "burst_write", "burst_read"],
    "AXI4":       ["write", "read", "burst_write", "burst_read"],
    "AXI4-Stream":["send", "recv"],
    "SPI":        ["write", "read", "transfer"],
    "I2C":        ["write", "read"],
    "UART":       ["send", "recv"],
    "TileLink":   ["get", "put", "atomic"],
}

# ── Utilities ─────────────────────────────────────────────────────────────────

def safe_write(path: str, content: str) -> bool:
    """Write file; skip silently if it already exists. Returns True if written."""
    if os.path.exists(path):
        print(f"  [SKIP] {path}  (already exists)")
        return False
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        f.write(content)
    print(f"  [GEN]  {path}")
    return True


def force_write(path: str, content: str):
    """Always write (for package files that must be regenerated)."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        f.write(content)
    print(f"  [GEN]  {path}")


def proto_to_vip(protocol: str) -> str:
    """Convert protocol name to vip prefix, e.g. 'AXI4-Stream' → 'axi4s_vip'."""
    mapping = {
        "APB":         "apb_vip",
        "AHB":         "ahb_vip",
        "AHB-Lite":    "ahb_vip",
        "AXI4":        "axi4_vip",
        "AXI4-Stream": "axi4s_vip",
        "SPI":         "spi_vip",
        "I2C":         "i2c_vip",
        "UART":        "uart_vip",
        "TileLink":    "tilelink_vip",
    }
    return mapping.get(protocol, protocol.lower().replace("-", "_") + "_vip")


def proto_prefix(protocol: str) -> str:
    """Short prefix for a protocol, e.g. 'AXI4-Stream' → 'axi4s'."""
    return proto_to_vip(protocol).replace("_vip", "")


def proto_dir_name(protocol: str) -> str:
    """Directory name for protocol sequences."""
    return protocol.upper().replace("-", "_").replace(" ", "_")


def extract_test_name(verification_type: str) -> tuple:
    """
    Parse verification_type cell.
    Returns (test_name, vtype) where vtype in ('directed','random','coverpoint','none').
    """
    if not verification_type:
        return None, "none"
    m = re.search(r'Testcase\(Directed\):\s*(\S+)', verification_type)
    if m:
        return m.group(1).strip(), "directed"
    m = re.search(r'Testcase\(Random\):\s*(\S+)', verification_type)
    if m:
        return m.group(1).strip(), "random"
    if "Coverpoint:" in verification_type:
        return None, "coverpoint"
    return None, "none"


def sanitize(name: str) -> str:
    """Make a string safe for use as SV identifier."""
    return re.sub(r'[^a-zA-Z0-9_]', '_', name).lower()


def parse_testcase_sections(desc: str) -> dict:
    """
    Parse structured testcase_description into sections.
    Expected keys: DUT Config, Stimulus, Expected Behavior, Checks, Pass Criteria, Notes
    """
    sections = {}
    current = None
    for line in desc.splitlines():
        line = line.strip()
        for key in ["DUT Config", "Stimulus", "Expected Behavior",
                    "Checks", "Pass Criteria", "Notes"]:
            if line.startswith(key + ":"):
                current = key
                sections[current] = line[len(key)+1:].strip()
                break
        else:
            if current and line:
                sections[current] = sections.get(current, "") + " " + line
    return sections


# ── Testplan parsers ──────────────────────────────────────────────────────────

def parse_testplan_json(path: str) -> list:
    """Parse testplan_data.json → list of row dicts."""
    with open(path) as f:
        data = json.load(f)
    rows = data.get("rows", [])
    result = []
    for r in rows:
        test_name, vtype = extract_test_name(r.get("verification_type", ""))
        if not test_name:
            continue
        result.append({
            "feature":              r.get("feature", ""),
            "subfeature":          r.get("subfeature", ""),
            "brief_description":   r.get("brief_description", ""),
            "verification_type":   vtype,
            "test_name":           test_name,
            "testcase_description":r.get("testcase_description", ""),
            "checker_id":          r.get("checker_id", ""),
            "checker_type":        r.get("checker_type", "Procedural"),
            "assertion_code":      r.get("assertion_code", ""),
            "milestone":           r.get("milestone", "DV-C"),
        })
    return result


def parse_testplan_excel(path: str) -> list:
    """Parse testplan.xlsx → list of row dicts."""
    if not HAS_OPENPYXL:
        print("WARNING: openpyxl not installed; cannot parse Excel. Install with: pip3 install openpyxl")
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Find the Testplan sheet (not Summary)
    sheet = None
    for name in wb.sheetnames:
        if name.lower() != "summary":
            sheet = wb[name]
            break
    if sheet is None:
        print("WARNING: Could not find Testplan sheet in workbook.")
        return []

    result = []
    header_seen = False
    for row in sheet.iter_rows(values_only=True):
        if not header_seen:
            # Skip until we see a row that looks like a header
            if row[0] and str(row[0]).strip().lower() in ("feature", "col 1"):
                header_seen = True
            continue
        if not row[0]:
            continue  # skip empty rows
        cols = [str(c).strip() if c is not None else "" for c in row]
        vtype_str = cols[3] if len(cols) > 3 else ""
        test_name, vtype = extract_test_name(vtype_str)
        if not test_name:
            continue
        result.append({
            "feature":              cols[0] if len(cols) > 0 else "",
            "subfeature":          cols[1] if len(cols) > 1 else "",
            "brief_description":   cols[2] if len(cols) > 2 else "",
            "verification_type":   vtype,
            "test_name":           test_name,
            "testcase_description":cols[4] if len(cols) > 4 else "",
            "checker_id":          cols[6] if len(cols) > 6 else "",
            "checker_type":        cols[7] if len(cols) > 7 else "Procedural",
            "assertion_code":      cols[8] if len(cols) > 8 else "",
            "milestone":           cols[9] if len(cols) > 9 else "DV-C",
        })
    return result


def load_testplan(input_data: dict) -> list:
    """Load testplan from input_data, preferring xlsx."""
    rows = input_data.get("testplan_rows")
    if rows:
        return rows
    xlsx = input_data.get("testplan_xlsx")
    if xlsx and os.path.exists(xlsx):
        print(f"  Parsing testplan Excel: {xlsx}")
        return parse_testplan_excel(xlsx)
    json_tp = input_data.get("testplan_json")
    if json_tp and os.path.exists(json_tp):
        print(f"  Parsing testplan JSON: {json_tp}")
        return parse_testplan_json(json_tp)
    print("WARNING: No testplan data found in input.")
    return []


# ── Agent sequence generators ─────────────────────────────────────────────────

def gen_agent_seq_file(proto: str, operation: str, vip_data: dict, dv_root: str) -> str:
    """Generate one agent sequence file. Returns file path."""
    proto_dir   = proto_dir_name(proto)
    prefix      = proto_prefix(proto)
    vip_name    = vip_data.get("vip_name", proto_to_vip(proto))
    data_width  = vip_data.get("data_width", 32)
    addr_width  = vip_data.get("addr_width", 32)
    signals     = PROTOCOL_SIGNALS.get(proto, [])

    class_name  = f"{prefix}_{operation}_seq"
    seq_dir     = os.path.join(dv_root, "sequences", proto_dir)
    file_path   = os.path.join(seq_dir, f"{class_name}.sv")

    # Build operation-specific randomize constraint
    op_constraint = _op_constraint(proto, operation)
    op_body       = _op_body(proto, operation)

    has_addr  = addr_width > 0
    has_data  = data_width > 0

    addr_field = f"  rand logic [{addr_width-1}:0] addr;\n" if has_addr else ""
    data_field = f"  rand logic [{data_width-1}:0] data;\n" if has_data else ""

    content = f"""\
// =============================================================================
// FILE: dv/sequences/{proto_dir}/{class_name}.sv
// S6 auto-generated — {date.today().isoformat()}
// Protocol : {proto}
// Operation: {operation}
// =============================================================================
class {class_name} extends {prefix}_base_seq;
  `uvm_object_utils({class_name})

{addr_field}{data_field}
  function new(string name = "{class_name}");
    super.new(name);
  endfunction

  task body();
    {prefix}_seq_item req;
    req = {prefix}_seq_item::type_id::create("req");
    start_item(req);
    if (!req.randomize() with {{
{op_constraint}
    }}) `uvm_fatal(`gtn, "{class_name}: randomization failed")
    finish_item(req);
{op_body}
  endtask
endclass
"""
    safe_write(file_path, content)
    return file_path


def _op_constraint(proto: str, operation: str) -> str:
    """Return inline randomize constraint for a given protocol/operation."""
    p = proto.upper().replace("-", "_")
    if "write" in operation or operation == "put" or operation == "send":
        if proto in ("APB",):
            return "      pwrite == 1'b1;"
        if proto in ("AHB", "AHB-Lite"):
            return "      hwrite == 1'b1;\n      htrans == 2'b10; // NONSEQ"
        if proto in ("AXI4",):
            return "      // AXI4 write — awvalid driven by driver"
        if proto in ("SPI",):
            return "      // SPI write transaction"
        if proto in ("UART",):
            return "      // UART send byte"
        return "      // TODO: add operation-specific constraints"
    if "read" in operation or operation == "get" or operation == "recv":
        if proto in ("APB",):
            return "      pwrite == 1'b0;"
        if proto in ("AHB", "AHB-Lite"):
            return "      hwrite == 1'b0;\n      htrans == 2'b10; // NONSEQ"
        return "      // TODO: add read constraints"
    if "burst" in operation:
        return "      // TODO: set burst length and type constraints"
    if operation in ("poll",):
        return "      pwrite == 1'b0; // poll = repeated read"
    return "      // TODO: add operation-specific constraints"


def _op_body(proto: str, operation: str) -> str:
    """Return post-finish_item body for a given protocol/operation."""
    lines = []
    lines.append("    // CHK: Embed Checker ID display when checker is associated.")
    lines.append("    // Example:")
    lines.append("    // `uvm_info(`gtn, \"[PASS] CHK_XXX — {proto} {operation} completed\", UVM_MEDIUM)")
    return "\n".join(lines)


def gen_all_agent_sequences(tb_data: dict, dv_root: str) -> dict:
    """
    Generate agent sequence files for every unique VIP.
    Returns dict: vip_name → list of generated file paths.
    """
    result = {}
    unique_vips = tb_data.get("unique_vips", [])
    for vip in unique_vips:
        proto    = vip.get("protocol", "")
        ops      = PROTOCOL_OPERATIONS.get(proto, ["write", "read"])
        files    = []
        for op in ops:
            fpath = gen_agent_seq_file(proto, op, vip, dv_root)
            files.append(fpath)
        result[vip.get("vip_name", proto_to_vip(proto))] = files
    return result


# ── Base vseq generator ───────────────────────────────────────────────────────

def gen_base_vseq(tb_data: dict, dv_root: str) -> str:
    """Generate <proj>_base_vseq.sv. Skips if exists."""
    proj         = tb_data.get("project_name", "proj")
    env_data     = tb_data.get("env", {})
    vseqr_class  = env_data.get("virtual_seqr_class", f"{proj}_virtual_seqr")
    ral_info     = tb_data.get("ral", {})
    ral_class    = ral_info.get("reg_block_class", f"{proj}_reg_block")

    seq_dir  = os.path.join(dv_root, "sequences", proj)
    filepath = os.path.join(seq_dir, f"{proj}_base_vseq.sv")

    content = f"""\
// =============================================================================
// FILE: dv/sequences/{proj}/{proj}_base_vseq.sv
// S6 auto-generated — {date.today().isoformat()}
// Base virtual sequence: RAL helpers, timing helpers, reset utility
// =============================================================================
class {proj}_base_vseq extends uvm_sequence #(uvm_sequence_item);
  `uvm_object_utils({proj}_base_vseq)
  `uvm_declare_p_sequencer({vseqr_class})

  // RAL handle — populated by base_test before vseq is started
  {ral_class} ral;

  function new(string name = "{proj}_base_vseq");
    super.new(name);
  endfunction

  // ── RAL helpers ─────────────────────────────────────────────────────────────
  task ral_write(uvm_reg rg, uvm_reg_data_t data, string chk_id = "");
    uvm_status_e status;
    rg.write(status, data, UVM_FRONTDOOR, .parent(this));
    if (status != UVM_IS_OK)
      `uvm_error(`gtn, $sformatf("[FAIL] %s RAL write to %s failed",
                                  chk_id, rg.get_name()))
    else if (chk_id != "")
      `uvm_info(`gtn, $sformatf("[PASS] %s — RAL write %s=0x%0h",
                                 chk_id, rg.get_name(), data), UVM_MEDIUM)
  endtask

  task ral_read(uvm_reg rg, output uvm_reg_data_t data, input string chk_id = "");
    uvm_status_e status;
    rg.read(status, data, UVM_FRONTDOOR, .parent(this));
    if (status != UVM_IS_OK)
      `uvm_error(`gtn, $sformatf("[FAIL] %s RAL read from %s failed",
                                  chk_id, rg.get_name()))
  endtask

  task ral_check(uvm_reg rg, uvm_reg_data_t exp_data, string chk_id = "");
    uvm_reg_data_t act_data;
    ral_read(rg, act_data);
    if (act_data !== exp_data)
      `uvm_error(`gtn, $sformatf("[FAIL] %s — %s: exp=0x%0h act=0x%0h",
                                  chk_id, rg.get_name(), exp_data, act_data))
    else
      `uvm_info(`gtn, $sformatf("[PASS] %s — %s matches 0x%0h",
                                 chk_id, rg.get_name(), exp_data), UVM_MEDIUM)
  endtask

  task ral_reset_walk(string chk_prefix = "CHK_RAL_RESET");
    uvm_reg regs[$];
    ral.get_registers(regs);
    foreach (regs[i]) begin
      uvm_reg_data_t reset_val = regs[i].get_reset();
      ral_check(regs[i], reset_val,
                $sformatf("%s_%s", chk_prefix, regs[i].get_name().toupper()));
    end
  endtask

  task ral_backdoor_check(uvm_reg rg, uvm_reg_data_t exp_data, string chk_id = "");
    uvm_reg_data_t act_data;
    uvm_status_e status;
    rg.peek(status, act_data);
    if (act_data !== exp_data)
      `uvm_error(`gtn, $sformatf("[FAIL] %s — %s (backdoor): exp=0x%0h act=0x%0h",
                                  chk_id, rg.get_name(), exp_data, act_data))
    else
      `uvm_info(`gtn, $sformatf("[PASS] %s — %s (backdoor) matches 0x%0h",
                                 chk_id, rg.get_name(), exp_data), UVM_MEDIUM)
  endtask

  // ── Timing helpers ───────────────────────────────────────────────────────────
  task wait_clocks(int unsigned n = 1);
    // TODO: replace with clocking-block based wait if vif is accessible
    #(n * 10ns);
  endtask

  task wait_for_reset_done();
    // TODO: connect to rst_n via p_sequencer.cfg.vif or env_cfg
    #100ns;
  endtask

  // ── Override in derived vseqs ────────────────────────────────────────────────
  virtual task body();
  endtask
endclass
"""
    safe_write(filepath, content)
    return filepath


# ── Directed vseq generator ───────────────────────────────────────────────────

def gen_directed_vseq(row: dict, tb_data: dict, dv_root: str) -> str:
    """Generate a directed virtual sequence for one testplan row."""
    proj        = tb_data.get("project_name", "proj")
    test_name   = sanitize(row["test_name"])
    feature     = row.get("feature", "")
    subfeature  = row.get("subfeature", "")
    milestone   = row.get("milestone", "DV-C")
    chk_id      = row.get("checker_id", "")
    chk_type    = row.get("checker_type", "Procedural")
    assert_code = row.get("assertion_code", "")
    tc_desc     = row.get("testcase_description", "")
    brief_desc  = row.get("brief_description", "")

    env_data    = tb_data.get("env", {})
    vseqr_class = env_data.get("virtual_seqr_class", f"{proj}_virtual_seqr")
    ral_info    = tb_data.get("ral", {})
    ral_class   = ral_info.get("reg_block_class", f"{proj}_reg_block")
    registers   = ral_info.get("registers", [])

    sections    = parse_testcase_sections(tc_desc)

    # Build body stubs from parsed sections
    dut_config_block  = _build_dut_config_block(sections.get("DUT Config", ""), registers, chk_id)
    stimulus_block    = _build_stimulus_block(sections.get("Stimulus", ""), tb_data, chk_id)
    check_block       = _build_check_block(sections.get("Checks", ""),
                                           sections.get("Pass Criteria", ""),
                                           registers, chk_id)
    assert_block      = _embed_assertion(assert_code, chk_id, chk_type)

    seq_dir  = os.path.join(dv_root, "sequences", proj)
    filepath = os.path.join(seq_dir, f"{test_name}_vseq.sv")

    content = f"""\
// =============================================================================
// FILE: dv/sequences/{proj}/{test_name}_vseq.sv
// S6 auto-generated — {date.today().isoformat()}
// Feature    : {feature}
// Sub-feature: {subfeature}
// Description: {brief_desc}
// Milestone  : {milestone}
// Checker ID : {chk_id}
// Checker Typ: {chk_type}
// =============================================================================
class {test_name}_vseq extends {proj}_base_vseq;
  `uvm_object_utils({test_name}_vseq)

  function new(string name = "{test_name}_vseq");
    super.new(name);
  endfunction

  task body();
    // ── DUT Configuration ───────────────────────────────────────────────────
{dut_config_block}

    // ── Stimulus ────────────────────────────────────────────────────────────
{stimulus_block}

    // ── Checks / Expected Behavior ──────────────────────────────────────────
{check_block}
{assert_block}
  endtask
endclass
"""
    safe_write(filepath, content)
    return filepath


def _build_dut_config_block(config_text: str, registers: list, chk_id: str) -> str:
    """Generate RAL write stubs from DUT Config section text."""
    lines = []
    if not config_text.strip():
        lines.append("    // TODO: configure DUT via RAL before stimulus  // ⚠️ NEEDS_REVIEW")
        return "\n".join(lines)

    # Try to detect register name mentions
    reg_names = [r.get("name", "") for r in registers]
    found_any = False
    for rname in reg_names:
        if rname.lower() in config_text.lower():
            lines.append(f"    ral_write(ral.{rname.lower()}, 32'h0, \"{chk_id}\");")
            lines.append(f"    // TODO: set correct value for {rname} per testcase")
            found_any = True
    if not found_any:
        lines.append(f"    // DUT Config: {config_text[:120]}")
        lines.append("    // TODO: translate above into RAL writes  // ⚠️ NEEDS_REVIEW")
    return "\n".join(lines)


def _build_stimulus_block(stim_text: str, tb_data: dict, chk_id: str) -> str:
    """Generate agent sequence start stubs from Stimulus section text."""
    lines = []
    if not stim_text.strip():
        lines.append("    // TODO: start agent sequences on p_sequencer  // ⚠️ NEEDS_REVIEW")
        return "\n".join(lines)

    proj = tb_data.get("project_name", "proj")
    unique_vips = tb_data.get("unique_vips", [])
    found_any = False
    for vip in unique_vips:
        proto  = vip.get("protocol", "")
        prefix = proto_prefix(proto)
        agent  = vip.get("agent_name", "")
        seqr   = vip.get("sequencer_name", f"m_{agent}_seqr") if agent else ""
        # Detect protocol mention in stimulus text
        if proto.lower() in stim_text.lower() or prefix in stim_text.lower():
            lines.append(f"    begin  // {proto} stimulus")
            lines.append(f"      {prefix}_write_seq {prefix}_seq;")
            lines.append(f"      {prefix}_seq = {prefix}_write_seq::type_id::create(\"{prefix}_seq\");")
            lines.append(f"      if (!{prefix}_seq.randomize())")
            lines.append(f"        `uvm_fatal(`gtn, \"{prefix}_write_seq randomize failed\")")
            lines.append(f"      {prefix}_seq.start(p_sequencer.{seqr});")
            lines.append(f"    end")
            found_any = True
    if not found_any:
        lines.append(f"    // Stimulus: {stim_text[:120]}")
        lines.append("    // TODO: start appropriate agent sequences  // ⚠️ NEEDS_REVIEW")
    return "\n".join(lines)


def _build_check_block(checks_text: str, pass_criteria: str, registers: list, chk_id: str) -> str:
    """Generate check stubs."""
    lines = []
    if not checks_text.strip() and not pass_criteria.strip():
        lines.append(f"    // TODO: implement checks for {chk_id}  // ⚠️ NEEDS_REVIEW")
        return "\n".join(lines)

    reg_names = [r.get("name", "") for r in registers]
    found_any = False
    for rname in reg_names:
        if rname.lower() in checks_text.lower() or rname.lower() in pass_criteria.lower():
            lines.append(f"    ral_check(ral.{rname.lower()}, 32'h0, \"{chk_id}\");")
            lines.append(f"    // TODO: set correct expected value for {rname}  // ⚠️ NEEDS_REVIEW")
            found_any = True
    if not found_any:
        if chk_id:
            lines.append(f"    // Pass Criteria: {pass_criteria[:100]}")
            lines.append(f"    // TODO: implement check below  // ⚠️ NEEDS_REVIEW")
            lines.append(f"    `uvm_info(`gtn, \"[PASS] {chk_id} — TODO: implement pass condition\", UVM_MEDIUM)")
    return "\n".join(lines)


def _embed_assertion(assert_code: str, chk_id: str, chk_type: str) -> str:
    """Embed SVA from testplan inline as a comment with pass action added."""
    if not assert_code.strip():
        return ""
    if chk_type not in ("Assertion", "Both"):
        return ""
    # Add pass action block if not present
    sva = assert_code.strip()
    if "assert property" in sva and "else" in sva and "$info" not in sva:
        sva = sva.replace(
            f'$error("[%s] {chk_id}',
            f'$info("[PASS] {chk_id} passed at %0t", $time);\n  $error("[%s] {chk_id}',
        )
    lines = []
    lines.append("    // ── Embedded SVA (from testplan) ─────────────────────────────────────")
    for line in sva.splitlines():
        lines.append(f"    // {line}")
    lines.append("    // ⚠️ NEEDS_REVIEW: instantiate this SVA in the interface or bind file")
    return "\n".join(lines) + "\n"


# ── Random vseq generator ─────────────────────────────────────────────────────

def gen_rand_vseq(row: dict, tb_data: dict, dv_root: str) -> str:
    """Generate a random virtual sequence for one testplan random row."""
    proj        = tb_data.get("project_name", "proj")
    test_name   = sanitize(row["test_name"])
    feature     = row.get("feature", "")
    milestone   = row.get("milestone", "DV-C")
    chk_id      = row.get("checker_id", "")
    brief_desc  = row.get("brief_description", "")

    env_data    = tb_data.get("env", {})
    vseqr_class = env_data.get("virtual_seqr_class", f"{proj}_virtual_seqr")
    ral_info    = tb_data.get("ral", {})
    ral_class   = ral_info.get("reg_block_class", f"{proj}_reg_block")

    unique_vips = tb_data.get("unique_vips", [])

    # Build agent sequence start blocks for all VIPs
    agent_blocks = []
    for vip in unique_vips:
        proto  = vip.get("protocol", "")
        prefix = proto_prefix(proto)
        agent  = vip.get("agent_name", "")
        seqr   = vip.get("sequencer_name", f"m_{agent}_seqr") if agent else f"m_{prefix}_seqr"
        agent_blocks.append(f"""\
      begin  // {proto}
        {prefix}_write_seq seq_{prefix};
        seq_{prefix} = {prefix}_write_seq::type_id::create("seq_{prefix}");
        if (!seq_{prefix}.randomize())
          `uvm_fatal(`gtn, "{prefix}_write_seq randomize failed")
        seq_{prefix}.start(p_sequencer.{seqr});
        // Randomness: seq_item fields randomized by {proto} driver constraints
      end""")

    agent_start = "\n".join(agent_blocks) if agent_blocks else \
        "        // TODO: start randomized sequences on each agent  // ⚠️ NEEDS_REVIEW"

    seq_dir  = os.path.join(dv_root, "sequences", proj)
    filepath = os.path.join(seq_dir, f"{test_name}_vseq.sv")

    content = f"""\
// =============================================================================
// FILE: dv/sequences/{proj}/{test_name}_vseq.sv
// S6 auto-generated — {date.today().isoformat()}
// Feature    : {feature}
// Description: {brief_desc}
// Milestone  : {milestone}
// Checker ID : {chk_id}
// Type       : RANDOM (randomness from vseq constraints + cfg + seq_item)
// =============================================================================
class {test_name}_vseq extends {proj}_base_vseq;
  `uvm_object_utils({test_name}_vseq)

  // ── Randomizable test parameters ──────────────────────────────────────────
  rand int unsigned num_txns;
  rand bit          inject_error;
  rand bit [1:0]    burst_type;

  constraint c_defaults {{
    num_txns    inside {{[10:200]}};
    inject_error dist  {{0 := 90, 1 := 10}};
    burst_type  inside {{2'b00, 2'b01, 2'b10}};
  }}

  // ── Apply test-specific constraint overrides in derived test (via plusarg or cfg) ──
  // Example override in test:
  //   vseq.randomize() with {{ num_txns == 500; inject_error == 1; }};

  function new(string name = "{test_name}_vseq");
    super.new(name);
  endfunction

  task body();
    // ── Optional: configure DUT mode before transactions ────────────────────
    // TODO: RAL writes for DUT configuration if needed  // ⚠️ NEEDS_REVIEW

    // ── Random stimulus loop ─────────────────────────────────────────────────
    repeat (num_txns) begin
      // Randomness sources:
      //   1. num_txns, inject_error, burst_type  (this vseq)
      //   2. seq_item fields                     (agent driver randomizes)
      //   3. env_cfg fields                      (constrained in test build_phase)
{agent_start}

      // Coverage sampling handled by VIP monitor → <vip>_coverage.sv
      // Checker ID display handled by scoreboard (S8) — embed here if procedural
      // `uvm_info(`gtn, "[PASS] {chk_id} — iteration done", UVM_HIGH)
    end

    // ── End-of-test checks ────────────────────────────────────────────────
    // TODO: final scoreboard / RAL state check after all transactions  // ⚠️ NEEDS_REVIEW
    if ("{chk_id}" != "")
      `uvm_info(`gtn, $sformatf("[PASS] {chk_id} — random test completed %0d txns",
                                 num_txns), UVM_MEDIUM)
  endtask
endclass
"""
    safe_write(filepath, content)
    return filepath


# ── Test class generator ──────────────────────────────────────────────────────

def gen_test_class(row: dict, tb_data: dict, dv_root: str) -> str:
    """Generate a UVM test class for one testplan row."""
    proj        = tb_data.get("project_name", "proj")
    test_name   = sanitize(row["test_name"])
    feature     = row.get("feature", "")
    subfeature  = row.get("subfeature", "")
    milestone   = row.get("milestone", "DV-C")
    chk_id      = row.get("checker_id", "")
    vtype       = row.get("verification_type", "directed")
    brief_desc  = row.get("brief_description", "")

    env_data    = tb_data.get("env", {})
    env_class   = env_data.get("env_class", f"{proj}_env")
    cfg_class   = env_data.get("env_cfg_class", f"{proj}_env_cfg")
    vseqr_class = env_data.get("virtual_seqr_class", f"{proj}_virtual_seqr")
    base_test   = tb_data.get("base_test", {}).get("base_test_class", f"{proj}_base_test")
    ral_info    = tb_data.get("ral", {})
    ral_class   = ral_info.get("reg_block_class", f"{proj}_reg_block")

    is_random   = (vtype == "random")
    vseq_class  = f"{test_name}_vseq"

    # Derive a feature-specific plusarg name
    feature_tag = sanitize(feature).upper()

    tests_dir = os.path.join(dv_root, "tests")
    filepath  = os.path.join(tests_dir, f"{proj}_{test_name}_test.sv")

    rand_cfg_block = ""
    if is_random:
        rand_cfg_block = f"""\
    // Random test: allow cfg to drive VIP active/passive and feature enables
    begin
      int num_txns_arg;
      if ($value$plusargs("NUM_TXNS=%d", num_txns_arg))
        cfg.num_txns = num_txns_arg;
    end
    // TODO: add test-specific cfg randomize-with constraints  // ⚠️ NEEDS_REVIEW
    // if (!cfg.randomize() with {{ num_txns inside {{[50:500]}}; }})
    //   `uvm_fatal(`gtn, "cfg randomization failed")
"""

    content = f"""\
// =============================================================================
// FILE: dv/tests/{proj}_{test_name}_test.sv
// S6 auto-generated — {date.today().isoformat()}
// Feature    : {feature}
// Sub-feature: {subfeature}
// Description: {brief_desc}
// Milestone  : {milestone}
// Checker ID : {chk_id}
// Type       : {'RANDOM' if is_random else 'DIRECTED'}
//
// Plusargs:
//   +VERBOSE              — UVM_HIGH verbosity
//   +NUM_TXNS=<N>         — override num_txns in vseq (random tests)
//   +SEED=<N>             — explicit randomization seed (set via simulator)
//   +SB_DISABLE           — disable scoreboard for this test
//   +COV_DISABLE          — disable functional coverage sampling
//   +{feature_tag}_MODE=<N> — feature-specific mode override
// =============================================================================
class {proj}_{test_name}_test extends {base_test};
  `uvm_component_utils({proj}_{test_name}_test)

  function new(string name = "{proj}_{test_name}_test", uvm_component parent = null);
    super.new(name, parent);
  endfunction

  // ── build_phase: apply test-specific cfg overrides ─────────────────────────
  function void build_phase(uvm_phase phase);
    super.build_phase(phase);

    // Plusarg overrides
    if ($test$plusargs("VERBOSE"))
      uvm_top.set_report_verbosity_level_hier(UVM_HIGH);
    if ($test$plusargs("SB_DISABLE"))
      cfg.sb_enable = 0;
    else
      cfg.sb_enable = 1;  // Default: scoreboard enabled for this test
    if ($test$plusargs("COV_DISABLE"))
      cfg.cov_enable = 0;

    begin
      int mode_arg;
      if ($value$plusargs("{feature_tag}_MODE=%d", mode_arg))
        cfg.{sanitize(feature)}_mode = mode_arg;
    end
{rand_cfg_block}
    // TODO: add any additional test-specific cfg settings  // ⚠️ NEEDS_REVIEW
  endfunction

  // ── run_phase: start vseq ──────────────────────────────────────────────────
  task run_phase(uvm_phase phase);
    {vseq_class} vseq;
    phase.raise_objection(this, "{proj}_{test_name}_test objection");

    vseq = {vseq_class}::type_id::create("vseq");
    vseq.ral = env.ral;

    if (!vseq.randomize())
      `uvm_fatal(`gtn, "{vseq_class} randomization failed")

    vseq.start(env.virtual_seqr);

    // Optional settling time after vseq
    #100ns;
    phase.drop_objection(this, "{proj}_{test_name}_test objection");
  endtask

  // ── report_phase: PASS/FAIL summary with Checker ID ───────────────────────
  function void report_phase(uvm_phase phase);
    uvm_report_server svr = uvm_report_server::get_server();
    super.report_phase(phase);
    if (svr.get_severity_count(UVM_ERROR) == 0 &&
        svr.get_severity_count(UVM_FATAL) == 0) begin
      `uvm_info(`gtn, $sformatf("[PASS] {chk_id} — %s completed with 0 errors",
                                 get_type_name()), UVM_NONE)
    end else begin
      `uvm_error(`gtn, $sformatf("[FAIL] {chk_id} — %s completed with %0d error(s)",
                  get_type_name(), svr.get_severity_count(UVM_ERROR)))
    end
  endfunction
endclass
"""
    safe_write(filepath, content)
    return filepath


# ── Package generators ────────────────────────────────────────────────────────

def gen_sequences_pkg(tb_data: dict, agent_seq_files: dict,
                      vseq_files: list, dv_root: str):
    """Generate <proj>_sequences_pkg.sv."""
    proj        = tb_data.get("project_name", "proj")
    unique_vips = tb_data.get("unique_vips", [])

    imports = "\n".join(
        f"  import {v.get('vip_name', proto_to_vip(v.get('protocol','')))}_pkg::*;"
        for v in unique_vips
    )

    # Agent sequence includes
    agent_includes = []
    for vip_name, files in agent_seq_files.items():
        for fpath in files:
            rel = os.path.relpath(fpath, dv_root)
            agent_includes.append(f'  `include "{rel}"')
    agent_block = "\n".join(agent_includes)

    # Vseq includes
    vseq_includes = []
    for fpath in vseq_files:
        rel = os.path.relpath(fpath, dv_root)
        vseq_includes.append(f'  `include "{rel}"')
    vseq_block = "\n".join(vseq_includes)

    content = f"""\
// =============================================================================
// FILE: dv/sequences/{proj}_sequences_pkg.sv
// S6 auto-generated — {date.today().isoformat()}
// =============================================================================
package {proj}_sequences_pkg;
  import uvm_pkg::*;
  `include "uvm_macros.svh"

  // VIP agent package imports
{imports}

  // ── Agent sequences (per protocol) ──────────────────────────────────────
{agent_block}

  // ── Virtual sequences (project-level) ───────────────────────────────────
{vseq_block}
endpackage : {proj}_sequences_pkg
"""
    filepath = os.path.join(dv_root, "sequences", f"{proj}_sequences_pkg.sv")
    force_write(filepath, content)


def gen_tests_pkg(tb_data: dict, test_files: list, dv_root: str):
    """Generate <proj>_tests_pkg.sv."""
    proj    = tb_data.get("project_name", "proj")
    env_pkg = f"{proj}_env_pkg"

    test_includes = []
    for fpath in test_files:
        rel = os.path.relpath(fpath, dv_root)
        test_includes.append(f'  `include "{rel}"')
    test_block = "\n".join(test_includes)

    content = f"""\
// =============================================================================
// FILE: dv/tests/{proj}_tests_pkg.sv
// S6 auto-generated — {date.today().isoformat()}
// =============================================================================
package {proj}_tests_pkg;
  import uvm_pkg::*;
  `include "uvm_macros.svh"
  import {env_pkg}::*;
  import {proj}_sequences_pkg::*;

  // ── Test classes ─────────────────────────────────────────────────────────
{test_block}
endpackage : {proj}_tests_pkg
"""
    filepath = os.path.join(dv_root, "tests", f"{proj}_tests_pkg.sv")
    force_write(filepath, content)


# ── compile.f update ──────────────────────────────────────────────────────────

def update_compile_f(dv_root: str, proj: str):
    """Append sequence/test package entries to compile.f if not already present."""
    compile_f = os.path.join(dv_root, "compile.f")
    seq_pkg   = f"${{DV_ROOT}}/sequences/{proj}_sequences_pkg.sv"
    test_pkg  = f"${{DV_ROOT}}/tests/{proj}_tests_pkg.sv"

    existing = ""
    if os.path.exists(compile_f):
        with open(compile_f) as f:
            existing = f.read()

    additions = []
    if seq_pkg not in existing:
        additions.append(f"+incdir+${{DV_ROOT}}/sequences")
        additions.append(seq_pkg)
    if test_pkg not in existing:
        additions.append(f"+incdir+${{DV_ROOT}}/tests")
        additions.append(test_pkg)

    if additions:
        with open(compile_f, "a") as f:
            f.write("\n// S6 sequences and tests\n")
            f.write("\n".join(additions) + "\n")
        print(f"  [UPD]  {compile_f}  (+{len(additions)} entries)")
    else:
        print(f"  [SKIP] {compile_f}  (entries already present)")


# ── dv_sequences_data.json ────────────────────────────────────────────────────

def write_sequences_data(tb_data: dict, rows: list,
                         agent_seq_files: dict,
                         vseq_files: list, test_files: list,
                         dv_root: str):
    proj    = tb_data.get("project_name", "proj")
    seq_pkg = os.path.join("dv", "sequences", f"{proj}_sequences_pkg.sv")
    tst_pkg = os.path.join("dv", "tests", f"{proj}_tests_pkg.sv")

    sequences = []
    for i, row in enumerate(rows):
        if i < len(vseq_files):
            sequences.append({
                "file":        os.path.relpath(vseq_files[i], os.path.dirname(dv_root)),
                "class":       sanitize(row["test_name"]) + "_vseq",
                "type":        row.get("verification_type", "directed"),
                "feature":     row.get("feature", ""),
                "milestone":   row.get("milestone", "DV-C"),
                "checker_ids": [row.get("checker_id", "")] if row.get("checker_id") else [],
            })

    tests = []
    for i, row in enumerate(rows):
        if i < len(test_files):
            tests.append({
                "file":      os.path.relpath(test_files[i], os.path.dirname(dv_root)),
                "class":     f"{proj}_{sanitize(row['test_name'])}_test",
                "vseq":      sanitize(row["test_name"]) + "_vseq",
                "plusargs":  ["+VERBOSE", "+NUM_TXNS=N", "+SB_DISABLE", "+COV_DISABLE"],
                "milestone": row.get("milestone", "DV-C"),
            })

    agent_seqs = [
        {"vip": vip, "files": [os.path.relpath(f, os.path.dirname(dv_root)) for f in files]}
        for vip, files in agent_seq_files.items()
    ]

    directed = sum(1 for r in rows if r.get("verification_type") == "directed")
    random   = sum(1 for r in rows if r.get("verification_type") == "random")
    neg      = sum(1 for r in rows if r.get("verification_type") not in ("directed", "random"))

    data = {
        "skill":          "dv-sequences",
        "version":        "1.0",
        "project_name":   proj,
        "generated_date": date.today().isoformat(),
        "sequences":      sequences,
        "tests":          tests,
        "agent_sequences":agent_seqs,
        "packages": {
            "sequences_pkg": seq_pkg,
            "tests_pkg":     tst_pkg,
        },
        "summary": {
            "total_tests":  len(rows),
            "directed":     directed,
            "random":       random,
            "neg_stress":   neg,
            "dv_i":  sum(1 for r in rows if r.get("milestone") == "DV-I"),
            "dv_c":  sum(1 for r in rows if r.get("milestone") == "DV-C"),
            "dv_f":  sum(1 for r in rows if r.get("milestone") == "DV-F"),
        }
    }

    out_path = os.path.join(dv_root, "dv_sequences_data.json")
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2)
    print(f"  [GEN]  {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="S6 dv-sequences: generate UVM sequences and test classes"
    )
    parser.add_argument("--input",   required=True,
                        help="Path to sequences_input.json (written by S6 SKILL.md Step 5a)")
    parser.add_argument("--output",  required=True,
                        help="Path to dv/ root directory")
    parser.add_argument("--filter",  default="all",
                        help="Filter: all | DV-I | DV-C | DV-F | <test_name>")
    parser.add_argument("--no-agent-seqs", action="store_true",
                        help="Skip agent sequence generation")
    parser.add_argument("--no-tests",      action="store_true",
                        help="Skip test class generation")
    args = parser.parse_args()

    dv_root = os.path.abspath(args.output)

    # Load input JSON
    with open(args.input) as f:
        input_data = json.load(f)

    tb_data = input_data.get("tb_data", {})
    proj    = tb_data.get("project_name", input_data.get("project_name", "proj"))
    if not tb_data.get("project_name"):
        tb_data["project_name"] = proj

    # Load testplan rows
    rows = load_testplan(input_data)
    if not rows:
        print("ERROR: No testplan rows found. Exiting.")
        sys.exit(1)

    # Apply filter
    filt = args.filter.strip()
    if filt != "all":
        if filt in ("DV-I", "DV-C", "DV-F"):
            rows = [r for r in rows if r.get("milestone") == filt]
        else:
            rows = [r for r in rows if sanitize(r.get("test_name","")) == sanitize(filt)]
        print(f"  Filter '{filt}': {len(rows)} rows selected")

    print(f"\n[S6] Generating sequences for project: {proj}")
    print(f"     dv_root  : {dv_root}")
    print(f"     testcases: {len(rows)}")
    print()

    # ── 1. Agent sequences ─────────────────────────────────────────────────
    agent_seq_files = {}
    if not args.no_agent_seqs:
        print("[Agent Sequences]")
        agent_seq_files = gen_all_agent_sequences(tb_data, dv_root)

    # ── 2. Base vseq ──────────────────────────────────────────────────────
    print("\n[Virtual Sequences]")
    base_vseq_path = gen_base_vseq(tb_data, dv_root)

    # ── 3. Per-testcase vseqs ─────────────────────────────────────────────
    vseq_files = [base_vseq_path]
    for row in rows:
        vtype = row.get("verification_type", "directed")
        if vtype == "random":
            fpath = gen_rand_vseq(row, tb_data, dv_root)
        else:
            fpath = gen_directed_vseq(row, tb_data, dv_root)
        vseq_files.append(fpath)

    # ── 4. Test classes ────────────────────────────────────────────────────
    test_files = []
    if not args.no_tests:
        print("\n[Test Classes]")
        for row in rows:
            fpath = gen_test_class(row, tb_data, dv_root)
            test_files.append(fpath)

    # ── 5. Packages ────────────────────────────────────────────────────────
    print("\n[Packages]")
    gen_sequences_pkg(tb_data, agent_seq_files, vseq_files, dv_root)
    gen_tests_pkg(tb_data, test_files, dv_root)

    # ── 6. compile.f ──────────────────────────────────────────────────────
    print("\n[compile.f]")
    update_compile_f(dv_root, proj)

    # ── 7. dv_sequences_data.json ─────────────────────────────────────────
    print("\n[Metadata]")
    write_sequences_data(tb_data, rows, agent_seq_files, vseq_files, test_files, dv_root)

    # ── Summary ───────────────────────────────────────────────────────────
    directed = sum(1 for r in rows if r.get("verification_type") == "directed")
    random   = sum(1 for r in rows if r.get("verification_type") == "random")
    dv_i  = sum(1 for r in rows if r.get("milestone") == "DV-I")
    dv_c  = sum(1 for r in rows if r.get("milestone") == "DV-C")
    dv_f  = sum(1 for r in rows if r.get("milestone") == "DV-F")
    n_agent = sum(len(v) for v in agent_seq_files.values())
    todo_count = 0  # could grep generated files, left as 0 for speed

    print(f"""
============================================================
  DV Sequences — Complete
  Project  : {proj}
  Output   : {dv_root}
============================================================
  Agent sequences : {n_agent} files  ({len(agent_seq_files)} protocols)
  Virtual sequences: {len(vseq_files)} files
  Test classes    : {len(test_files)} files
------------------------------------------------------------
  Directed: {directed}   Random: {random}   Other: {len(rows)-directed-random}
  Milestone: DV-I={dv_i}  DV-C={dv_c}  DV-F={dv_f}
------------------------------------------------------------
  Packages:
    sequences_pkg: dv/sequences/{proj}_sequences_pkg.sv
    tests_pkg    : dv/tests/{proj}_tests_pkg.sv
------------------------------------------------------------
  ⚠️  Check TODOs: grep -r "NEEDS_REVIEW" {dv_root}/sequences {dv_root}/tests
------------------------------------------------------------
  Next step: /dv-assertions (S7) or /dv-scoreboard (S8)
============================================================
""")


if __name__ == "__main__":
    main()
