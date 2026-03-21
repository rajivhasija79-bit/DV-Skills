#!/usr/bin/env python3
"""Generate AXI DMA DV Testplan Excel."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# DATA
# ─────────────────────────────────────────────

PROJECT = "axi_dma"
BLOCK   = "AXI_DMA"
SPEC    = "axi_dma_spec.md v0.9"

# 11 columns for Testplan sheet
TP_HEADERS = [
    "Test ID",
    "Feature",
    "Sub-Feature",
    "Test Name",
    "Test Description",
    "Stimulus / Scenario",
    "Expected Behavior / Check",
    "Coverage Points",
    "Priority",
    "Test Type",
    "Status",
]

# (test_id, feature, sub_feature, test_name, description,
#  stimulus, expected, coverage, priority, test_type, status)
TESTPLAN = [
    # ── RESET ──────────────────────────────────────────────────────────────
    (
        "TC_RST_001",
        "Reset",
        "AXI Reset (aresetn)",
        "axi_async_reset",
        "Assert active-low AXI reset asynchronously and verify all AXI master outputs deassert.",
        "Assert aresetn=0 while aclk is running; hold for 10 cycles then deassert.",
        "awvalid, wvalid, arvalid, bready, rready all 0 during reset; DMA state=Idle after deassertion.",
        "reset_coverage: aresetn toggled; state == Idle post-reset",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_RST_002",
        "Reset",
        "APB Reset (presetn)",
        "apb_sync_reset",
        "Assert APB synchronous reset and verify all registers return to reset values.",
        "Assert presetn=0 for 5 pclk cycles; read all config registers after deassertion.",
        "SRC_ADDR=0, DST_ADDR=0, LENGTH=0, CTRL=0, STATUS=0 after reset.",
        "reset_coverage: presetn toggled; all register reset values sampled",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_RST_003",
        "Reset",
        "Reset During Active Transfer",
        "reset_mid_transfer",
        "Assert aresetn while a DMA transfer is in-flight and verify clean abort.",
        "Start a 256-byte transfer; after 3 AXI beats assert aresetn; monitor bus.",
        "AXI outputs deassert within 1 cycle; no response protocol violations; state=Idle after reset.",
        "FSM transition: Active -> Idle (via reset); reset_during_active_cp",
        "High",
        "Directed",
        "Not Started",
    ),

    # ── APB REGISTER ACCESS ────────────────────────────────────────────────
    (
        "TC_REG_001",
        "APB Register Access",
        "Register Write",
        "apb_reg_write_all",
        "Write legal values to all writable registers (SRC_ADDR, DST_ADDR, LENGTH, CTRL) and read back.",
        "APB write then APB read to each register with max and typical values.",
        "Read-back data matches written value for all writable fields; read-only fields unchanged.",
        "reg_write_cp: {SRC_ADDR, DST_ADDR, LENGTH, CTRL} written; readback verified",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_REG_002",
        "APB Register Access",
        "Register Read (STATUS)",
        "apb_status_read",
        "Read STATUS register in Idle, Active, and Error states; verify field accuracy.",
        "Read STATUS after reset, during transfer, and after injecting AXI error.",
        "STATUS.BUSY=0 in Idle; STATUS.BUSY=1 during transfer; STATUS.ERR=1 after error.",
        "status_reg_cp: {Idle, Active, Error} states sampled during STATUS read",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_REG_003",
        "APB Register Access",
        "Write to Read-Only Fields",
        "apb_ro_field_write",
        "Attempt to write read-only STATUS fields; confirm no state change.",
        "APB write to STATUS register address with non-zero data.",
        "STATUS register value unchanged; no side effects; pready asserted normally.",
        "ro_field_write_cp",
        "Medium",
        "Directed",
        "Not Started",
    ),
    (
        "TC_REG_004",
        "APB Register Access",
        "APB Timing — Setup/Enable phases",
        "apb_protocol_phases",
        "Verify correct pready handshake and two-phase APB access (Setup then Enable).",
        "Drive valid APB setup+enable cycles; check pready assertion.",
        "pready=1 on correct enable cycle; prdata stable; no protocol violation.",
        "apb_phase_cp: setup_only vs setup+enable phases covered",
        "Medium",
        "Directed",
        "Not Started",
    ),

    # ── TRANSFER — BASIC ───────────────────────────────────────────────────
    (
        "TC_XFR_001",
        "Memory-to-Memory Transfer",
        "Single-Beat Transfer",
        "single_beat_xfr",
        "Configure and start a 4-byte (single beat) DMA transfer; verify data movement.",
        "Write SRC_ADDR, DST_ADDR, LENGTH=4, CTRL.START=1; provide AXI slave response.",
        "AXI read then write for 1 beat; dma_done=1 after completion; STATUS.BUSY=0.",
        "xfr_length_cp: length==1 beat; FSM: Idle->Active->Idle",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_XFR_002",
        "Memory-to-Memory Transfer",
        "Max Length Transfer (65535 bytes)",
        "max_length_xfr",
        "Transfer maximum configurable length of 65535 bytes and verify correct beat count.",
        "Set LENGTH=65535; monitor AXI arlen/awlen for correct burst segmentation.",
        "All bytes transferred; data integrity preserved; dma_done=1; no protocol errors.",
        "xfr_length_cp: length==65535; max_length_cp",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_XFR_003",
        "Memory-to-Memory Transfer",
        "Minimum Length Transfer (1 byte)",
        "min_length_xfr",
        "Transfer 1 byte; verify wstrb reflects byte enable correctly.",
        "Set LENGTH=1; check wstrb on AXI write.",
        "Correct single-byte wstrb; dma_done=1; no extra beats generated.",
        "xfr_length_cp: length==1 byte; wstrb_cp",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_XFR_004",
        "Memory-to-Memory Transfer",
        "Word-Aligned Addresses",
        "word_aligned_xfr",
        "Verify transfer with word-aligned source and destination addresses.",
        "Use SRC_ADDR and DST_ADDR both word-aligned (addr[1:0]==2'b00).",
        "AXI araddr and awaddr are word-aligned; no address errors.",
        "addr_align_cp: aligned addresses",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_XFR_005",
        "Memory-to-Memory Transfer",
        "Data Integrity Check",
        "data_integrity_xfr",
        "Verify data read from source matches data written to destination across a full transfer.",
        "Randomize source memory content; run DMA; compare destination to source.",
        "Byte-for-byte match between source data and destination data written via AXI.",
        "data_integrity_cp",
        "High",
        "Constrained-Random",
        "Not Started",
    ),

    # ── BURST TRANSFERS ────────────────────────────────────────────────────
    (
        "TC_BST_001",
        "Burst Transfers",
        "Burst Length Minimum (1 beat)",
        "burst_len_1",
        "Issue burst transfer with burst length 1 (INCR, single beat).",
        "Configure MAX_BURST equivalent to 1; trigger transfer.",
        "arlen=0, awlen=0; single-beat AXI transactions; correct INCR burst type.",
        "burst_len_cp: len==1",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_BST_002",
        "Burst Transfers",
        "Burst Length Maximum (16 beats)",
        "burst_len_16",
        "Issue burst transfer with maximum burst length of 16 beats.",
        "Configure burst_len=16; check arlen=15, awlen=15.",
        "arlen=15, awlen=15; wlast asserts on beat 16; correct byte count.",
        "burst_len_cp: len==16; wlast_cp",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_BST_003",
        "Burst Transfers",
        "Mid-Range Burst Lengths",
        "burst_len_sweep",
        "Sweep burst lengths 2–15 and verify correctness for each.",
        "Constrained-random burst length in [2,15]; run transfer for each.",
        "Correct arlen/awlen for each length; wlast on correct beat; data integrity maintained.",
        "burst_len_cp: len in {2..15}",
        "Medium",
        "Constrained-Random",
        "Not Started",
    ),
    (
        "TC_BST_004",
        "Burst Transfers",
        "INCR Burst Type Assertion",
        "burst_type_incr",
        "Verify that AXI AxBURST is always INCR (2'b01) regardless of transfer parameters.",
        "Run multiple transfers with varying lengths; check AxBURST on every transaction.",
        "AxBURST == 2'b01 for all read and write address phases.",
        "burst_type_cp: INCR observed",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_BST_005",
        "Burst Transfers",
        "Transfer Spanning Multiple Bursts",
        "multi_burst_xfr",
        "Transfer a length requiring multiple bursts (e.g., 512 bytes at burst=16x4B).",
        "Set LENGTH=512, burst_len=16; count number of AXI AR/AW transactions.",
        "8 read bursts and 8 write bursts; addresses increment correctly; no data loss.",
        "multi_burst_cp; burst_count_cp",
        "High",
        "Directed",
        "Not Started",
    ),

    # ── AXI HANDSHAKE ──────────────────────────────────────────────────────
    (
        "TC_AXI_001",
        "AXI4 Protocol",
        "Address Channel Handshake",
        "axi_addr_handshake",
        "Verify arvalid/arready and awvalid/awready handshakes with variable ready delays.",
        "Randomize arready/awready deassertion for 0–10 cycles; check no data before addr accepted.",
        "DMA waits for ready; valid held stable until accepted; no address changes during valid.",
        "axi_addr_handshake_cp: {ready_same_cycle, ready_delayed} for read and write",
        "High",
        "Constrained-Random",
        "Not Started",
    ),
    (
        "TC_AXI_002",
        "AXI4 Protocol",
        "Write Data Channel Handshake",
        "axi_wdata_handshake",
        "Verify wvalid/wready handshake and wlast on final beat.",
        "Randomize wready deassertion; monitor wlast position.",
        "wlast=1 on last beat only; wvalid held until wready; wstrb correct per beat.",
        "wdata_handshake_cp; wlast_cp",
        "High",
        "Constrained-Random",
        "Not Started",
    ),
    (
        "TC_AXI_003",
        "AXI4 Protocol",
        "Write Response Channel",
        "axi_bresp_ok",
        "Verify DMA asserts bready and accepts OKAY write response correctly.",
        "Provide bresp=OKAY with random bvalid delay; check bready.",
        "bready=1; DMA proceeds after bvalid+bready handshake; STATUS.BUSY reflects completion.",
        "bresp_cp: OKAY response accepted",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_AXI_004",
        "AXI4 Protocol",
        "Read Data Channel Handshake",
        "axi_rdata_handshake",
        "Verify rready handshake and rlast reception.",
        "Randomize rvalid assertion delay; check rready and rlast handling.",
        "rready=1 when DMA ready; rlast detected to end read burst; data latched correctly.",
        "rdata_handshake_cp; rlast_cp",
        "High",
        "Constrained-Random",
        "Not Started",
    ),
    (
        "TC_AXI_005",
        "AXI4 Protocol",
        "Single Outstanding Transaction",
        "single_outstanding_xfr",
        "Confirm DMA issues only one outstanding transaction at a time (no pipelining).",
        "Monitor AXI channels; check no new AR/AW issued before previous response received.",
        "No overlapping AR and AW transactions; one outstanding read or write at a time.",
        "outstanding_xfr_cp",
        "High",
        "Directed",
        "Not Started",
    ),

    # ── INTERRUPTS ─────────────────────────────────────────────────────────
    (
        "TC_INT_001",
        "Interrupts",
        "Transfer Complete (dma_done)",
        "dma_done_interrupt",
        "Verify dma_done pulses high for one cycle upon successful transfer completion.",
        "Run a complete transfer; observe dma_done timing relative to final bresp.",
        "dma_done=1 for exactly 1 cycle after last write response; then 0.",
        "dma_done_cp: pulse width==1; timing relative to bresp",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_INT_002",
        "Interrupts",
        "Error Interrupt (dma_err)",
        "dma_err_interrupt",
        "Verify dma_err asserts when AXI slave returns SLVERR or DECERR.",
        "Inject bresp=SLVERR or rresp=SLVERR during active transfer.",
        "dma_err=1; DMA transitions to Error state; transfer halted.",
        "dma_err_cp: SLVERR; FSM: Active->Error",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_INT_003",
        "Interrupts",
        "No Spurious Interrupts",
        "no_spurious_int",
        "Verify dma_done and dma_err are not asserted spuriously during idle or mid-transfer.",
        "Monitor both interrupt lines for entire simulation including reset, idle, and mid-transfer.",
        "dma_done and dma_err remain 0 except at defined events.",
        "spurious_int_cp",
        "Medium",
        "Directed",
        "Not Started",
    ),

    # ── ERROR HANDLING ─────────────────────────────────────────────────────
    (
        "TC_ERR_001",
        "Error Handling",
        "AXI Write SLVERR",
        "axi_write_slverr",
        "Inject SLVERR on write response channel; verify DMA error behavior.",
        "Return bresp=2'b10 (SLVERR) on write transaction.",
        "DMA asserts dma_err; enters Error state; no further AXI transactions issued.",
        "error_cp: write_slverr; FSM: Active->Error",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_ERR_002",
        "Error Handling",
        "AXI Read SLVERR",
        "axi_read_slverr",
        "Inject SLVERR on read response channel during read phase.",
        "Return rresp=2'b10 (SLVERR) on read transaction.",
        "DMA asserts dma_err; enters Error state; write phase not started.",
        "error_cp: read_slverr; FSM: Active->Error",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_ERR_003",
        "Error Handling",
        "AXI Write DECERR",
        "axi_write_decerr",
        "Inject DECERR on write response; verify same error behavior as SLVERR.",
        "Return bresp=2'b11 (DECERR) on write transaction.",
        "DMA asserts dma_err; enters Error state; STATUS.ERR=1.",
        "error_cp: write_decerr; FSM: Active->Error",
        "Medium",
        "Directed",
        "Not Started",
    ),
    (
        "TC_ERR_004",
        "Error Handling",
        "AXI Read DECERR",
        "axi_read_decerr",
        "Inject DECERR on read response; verify error state entry.",
        "Return rresp=2'b11 (DECERR) on read transaction.",
        "DMA asserts dma_err; enters Error state; STATUS.ERR=1.",
        "error_cp: read_decerr; FSM: Active->Error",
        "Medium",
        "Directed",
        "Not Started",
    ),
    (
        "TC_ERR_005",
        "Error Handling",
        "Error State Software Clear",
        "error_state_sw_clear",
        "Clear DMA from Error state via software CTRL.CLR_ERR write; verify return to Idle.",
        "Enter Error state; write CTRL.CLR_ERR=1 via APB; observe state and STATUS.",
        "DMA transitions Error->Idle; STATUS.ERR=0; dma_err=0; ready for new transfer.",
        "FSM transition: Error->Idle; error_clear_cp",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_ERR_006",
        "Error Handling",
        "New Transfer Blocked in Error State",
        "new_xfr_blocked_in_error",
        "Attempt to start a new transfer while DMA is in Error state; verify blocked.",
        "Enter Error state; write CTRL.START=1 without clearing error first.",
        "DMA does not start new transfer; STATUS.BUSY=0; dma_err remains 1.",
        "error_block_cp",
        "Medium",
        "Directed",
        "Not Started",
    ),

    # ── FSM TRANSITION COVERAGE ───────────────────────────────────────────
    (
        "TC_FSM_001",
        "FSM Transitions",
        "Idle -> Active",
        "fsm_idle_to_active",
        "Verify DMA transitions from Idle to Active when CTRL.START is written.",
        "Write valid SRC_ADDR, DST_ADDR, LENGTH; then write CTRL.START=1.",
        "DMA state changes from Idle to Active; AXI AR transaction initiated.",
        "fsm_cp: Idle->Active transition",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_FSM_002",
        "FSM Transitions",
        "Active -> Idle (Completion)",
        "fsm_active_to_idle_ok",
        "Verify DMA returns to Idle after successful transfer completion.",
        "Complete full transfer with OKAY responses; monitor state after dma_done.",
        "State=Idle after dma_done pulse; STATUS.BUSY=0; ready for new transfer.",
        "fsm_cp: Active->Idle (success)",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_FSM_003",
        "FSM Transitions",
        "Active -> Error (AXI Error)",
        "fsm_active_to_error",
        "Verify DMA transitions from Active to Error on AXI error response.",
        "Inject SLVERR during active transfer.",
        "State=Error; dma_err=1; AXI transactions halted.",
        "fsm_cp: Active->Error transition",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_FSM_004",
        "FSM Transitions",
        "Error -> Idle (SW Clear)",
        "fsm_error_to_idle",
        "Verify DMA transitions from Error to Idle via software clear.",
        "Enter Error state; write CTRL.CLR_ERR=1.",
        "State=Idle; STATUS.ERR=0; DMA ready.",
        "fsm_cp: Error->Idle transition",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_FSM_005",
        "FSM Transitions",
        "Idle -> Active -> Idle (Back-to-Back)",
        "fsm_back_to_back_xfr",
        "Issue two consecutive transfers without any gap; verify clean state transitions.",
        "Complete transfer 1; immediately write CTRL.START=1 for transfer 2.",
        "Both transfers complete correctly; FSM: Idle->Active->Idle->Active->Idle.",
        "fsm_cp: consecutive Idle->Active->Idle; back_to_back_cp",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_FSM_006",
        "FSM Transitions",
        "Idle -> Active -> Error -> Idle (Full Error Cycle)",
        "fsm_full_error_cycle",
        "Exercise complete error cycle: start transfer, inject error, clear error, restart.",
        "Start transfer; inject SLVERR; clear error; start new transfer.",
        "Full FSM path exercised; second transfer completes successfully.",
        "fsm_cp: full error cycle",
        "High",
        "Directed",
        "Not Started",
    ),
    (
        "TC_FSM_007",
        "FSM Transitions",
        "Idle -> Idle (No-start)",
        "fsm_idle_stay",
        "Verify DMA remains in Idle when CTRL.START is not set.",
        "Write SRC/DST/LENGTH but do not write CTRL.START; check state.",
        "State remains Idle; no AXI activity.",
        "fsm_cp: Idle stays Idle without START",
        "Low",
        "Directed",
        "Not Started",
    ),
    (
        "TC_FSM_008",
        "FSM Transitions",
        "Active -> Idle (Reset)",
        "fsm_active_to_idle_reset",
        "Verify aresetn forces Active->Idle regardless of bus state.",
        "Assert aresetn during Active state.",
        "State=Idle immediately; AXI outputs deassert.",
        "fsm_cp: Active->Idle via reset",
        "High",
        "Directed",
        "Not Started",
    ),

    # ── ADDRESS / ALIGNMENT ───────────────────────────────────────────────
    (
        "TC_ADDR_001",
        "Address Handling",
        "Source Address Range",
        "src_addr_range",
        "Sweep source addresses across full 32-bit range (word-aligned) for corner cases.",
        "Use SRC_ADDR at 0x0, 0x4, 0xFFFFFFFC; run single-beat transfer.",
        "Correct araddr on AXI; no address wrapping or truncation.",
        "src_addr_cp: {min, mid, max} word-aligned addresses",
        "Medium",
        "Constrained-Random",
        "Not Started",
    ),
    (
        "TC_ADDR_002",
        "Address Handling",
        "Destination Address Range",
        "dst_addr_range",
        "Sweep destination addresses across 32-bit range (word-aligned).",
        "Use DST_ADDR at 0x0, 0x4, 0xFFFFFFFC; run single-beat transfer.",
        "Correct awaddr on AXI; no address wrapping or truncation.",
        "dst_addr_cp: {min, mid, max} word-aligned addresses",
        "Medium",
        "Constrained-Random",
        "Not Started",
    ),
    (
        "TC_ADDR_003",
        "Address Handling",
        "Address Increment Across Bursts",
        "addr_increment_across_bursts",
        "Verify address increments by correct byte offset for each burst and beat.",
        "Run multi-burst transfer; record araddr and awaddr for each transaction.",
        "Each burst address = previous burst address + (burst_len * 4); no gaps.",
        "addr_increment_cp",
        "High",
        "Directed",
        "Not Started",
    ),

    # ── CONSTRAINED RANDOM / REGRESSION ───────────────────────────────────
    (
        "TC_RND_001",
        "Constrained Random",
        "Random Transfer Parameters",
        "rand_xfr_params",
        "Randomize SRC_ADDR, DST_ADDR, LENGTH, burst_len and run many transfers.",
        "Constrained-random: addr word-aligned, length in [1,65535], burst in [1,16].",
        "All transfers complete; data integrity holds; no protocol violations.",
        "rand_xfr_cp: full parameter space sampled",
        "High",
        "Constrained-Random",
        "Not Started",
    ),
    (
        "TC_RND_002",
        "Constrained Random",
        "Random APB Interleaving",
        "rand_apb_during_xfr",
        "Randomly read STATUS register during active transfer; check no read side-effects.",
        "Interleave APB STATUS reads with random timing during AXI transfer.",
        "STATUS.BUSY=1 during transfer; no functional impact from APB reads.",
        "apb_interleave_cp",
        "Medium",
        "Constrained-Random",
        "Not Started",
    ),
    (
        "TC_RND_003",
        "Constrained Random",
        "Random AXI Backpressure",
        "rand_axi_backpressure",
        "Apply random backpressure on all AXI channels simultaneously during a transfer.",
        "Randomize arready, rvalid, awready, wready, bvalid assertion delays independently.",
        "Transfer completes correctly despite backpressure; no timeout or data corruption.",
        "backpressure_cp: all channels covered",
        "High",
        "Constrained-Random",
        "Not Started",
    ),

    # ── CLOCK/RESET EDGE CASES ────────────────────────────────────────────
    (
        "TC_CLK_001",
        "Clock & Reset",
        "AXI Clock Frequency Boundary",
        "axi_clk_freq_bound",
        "Run transfers at min (100 MHz) and max (200 MHz) AXI clock frequencies.",
        "Change aclk period; run full transfer at each boundary.",
        "Transfers complete without errors at both clock extremes.",
        "axi_clk_freq_cp: {100MHz, 200MHz}",
        "Medium",
        "Directed",
        "Not Started",
    ),
    (
        "TC_CLK_002",
        "Clock & Reset",
        "APB Clock Frequency Boundary",
        "apb_clk_freq_bound",
        "Run register accesses at min (50 MHz) and max (100 MHz) APB clock frequencies.",
        "Change pclk period; perform write/read register access.",
        "Register accesses succeed without errors at both clock extremes.",
        "apb_clk_freq_cp: {50MHz, 100MHz}",
        "Medium",
        "Directed",
        "Not Started",
    ),
]

# ─────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FONT   = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
TITLE_FONT    = Font(name="Calibri", bold=True, size=14, color="1F3864")
LABEL_FONT    = Font(name="Calibri", bold=True, size=11, color="1F3864")
BODY_FONT     = Font(name="Calibri", size=10)

HEADER_FILL   = make_fill("1F3864")   # dark navy
ALT_FILL      = make_fill("DCE6F1")   # light blue
WHITE_FILL    = make_fill("FFFFFF")

THIN_BORDER   = make_border("thin")

PRIORITY_FILL = {
    "High":   make_fill("FFD7D7"),
    "Medium": make_fill("FFFACC"),
    "Low":    make_fill("D7F0D7"),
}

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def style_header_cell(cell, text):
    cell.value     = text
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = THIN_BORDER


def style_body_cell(cell, text, alt=False, center=False):
    cell.value     = text
    cell.font      = BODY_FONT
    cell.fill      = ALT_FILL if alt else WHITE_FILL
    cell.alignment = CENTER if center else LEFT
    cell.border    = THIN_BORDER


# ─────────────────────────────────────────────
# SUMMARY SHEET
# ─────────────────────────────────────────────

def build_summary(wb):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value     = "AXI DMA — DV Testplan Summary"
    c.font      = TITLE_FONT
    c.alignment = CENTER

    # Meta table
    meta = [
        ("Project",        PROJECT),
        ("Block",          BLOCK),
        ("Spec Reference", SPEC),
        ("Total Tests",    str(len(TESTPLAN))),
        ("Date Generated", "2026-03-22"),
        ("Author",         "DV Automation"),
        ("Status",         "Draft"),
    ]
    for i, (label, value) in enumerate(meta, start=4):
        row = i
        lc = ws.cell(row=row, column=2, value=label)
        lc.font      = LABEL_FONT
        lc.alignment = LEFT
        lc.border    = THIN_BORDER
        lc.fill      = make_fill("D9E1F2")

        vc = ws.cell(row=row, column=3, value=value)
        vc.font      = BODY_FONT
        vc.alignment = LEFT
        vc.border    = THIN_BORDER
        vc.fill      = WHITE_FILL

    # Feature summary table header
    start_row = 4 + len(meta) + 2
    headers = ["Feature Area", "Test Count", "Priority Breakdown (H/M/L)"]
    for col, h in enumerate(headers, start=2):
        style_header_cell(ws.cell(row=start_row, column=col), h)

    # Gather feature stats
    from collections import Counter
    feat_counts  = Counter()
    feat_priority = {}
    for row in TESTPLAN:
        feat   = row[1]
        pri    = row[8]
        feat_counts[feat] += 1
        if feat not in feat_priority:
            feat_priority[feat] = Counter()
        feat_priority[feat][pri] += 1

    for i, (feat, cnt) in enumerate(sorted(feat_counts.items()), start=1):
        r    = start_row + i
        alt  = (i % 2 == 0)
        pc   = feat_priority[feat]
        breakdown = f"H={pc.get('High',0)} / M={pc.get('Medium',0)} / L={pc.get('Low',0)}"
        style_body_cell(ws.cell(row=r, column=2), feat,      alt)
        style_body_cell(ws.cell(row=r, column=3), str(cnt),  alt, center=True)
        style_body_cell(ws.cell(row=r, column=4), breakdown, alt, center=True)

    # Totals row
    last_r = start_row + len(feat_counts) + 1
    tc = ws.cell(row=last_r, column=2, value="TOTAL")
    tc.font      = Font(name="Calibri", bold=True, size=10)
    tc.fill      = make_fill("D9E1F2")
    tc.border    = THIN_BORDER
    tc.alignment = CENTER

    tn = ws.cell(row=last_r, column=3, value=len(TESTPLAN))
    tn.font      = Font(name="Calibri", bold=True, size=10)
    tn.fill      = make_fill("D9E1F2")
    tn.border    = THIN_BORDER
    tn.alignment = CENTER

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 30


# ─────────────────────────────────────────────
# TESTPLAN SHEET
# ─────────────────────────────────────────────

def build_testplan(wb):
    ws = wb.create_sheet("Testplan")
    ws.sheet_view.showGridLines = False

    # Column widths (11 columns)
    col_widths = [14, 26, 26, 28, 46, 46, 46, 36, 10, 20, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Header row
    for col, hdr in enumerate(TP_HEADERS, start=1):
        style_header_cell(ws.cell(row=1, column=col), hdr)
    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, entry in enumerate(TESTPLAN, start=2):
        alt = (row_idx % 2 == 0)
        for col_idx, value in enumerate(entry, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            center_cols = {1, 9, 10, 11}   # Test ID, Priority, Type, Status
            style_body_cell(cell, value, alt=alt, center=(col_idx in center_cols))

            # Priority colour on Priority column (col 9)
            if col_idx == 9 and value in PRIORITY_FILL:
                cell.fill = PRIORITY_FILL[value]

        ws.row_dimensions[row_idx].height = 48

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(TP_HEADERS))}1"


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    output_path = (
        "/Users/apple/Documents/DV-Skills/skills/dv-testplan/evals/workspace/"
        "iteration-1/eval-2-axi-dma-testplan/without_skill/outputs/testplan.xlsx"
    )

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_summary(wb)
    build_testplan(wb)

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Total test cases: {len(TESTPLAN)}")
